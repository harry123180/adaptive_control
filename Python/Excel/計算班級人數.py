from openpyxl import load_workbook
from matplotlib import pyplot as plt
import numpy as np
wb = load_workbook('C:\\Users\\user\\Desktop\\研究\\研究紀錄By泓舉\\7碩一上第七周\\cncdata2.xlsx')
sheet = wb['工作表2']
ws1 = wb.create_sheet("新增工作表1")
# 根據位置取得儲存格

single =0
Y_current = sheet['F26':'F63121']
x_data=[]
#7355
#7380
indx = 0
y_data = []
for row in Y_current:
    for c in row:
        #print(c.value)
        ws1['A'+str(indx+2)].value = c.value
        indx+=1
        y_data.append(c.value)
ws1['A1'].value = 'Y軸電流'
for i in range(len(y_data)):
    x_data.append(i)
print(len(x_data),len(y_data))

#print(ws1['A1'].value )
list_len = len(x_data)
#wb.save('C:\\Users\\user\\Desktop\\研究\\研究紀錄By泓舉\\7碩一上第七周\\cncdata1.xlsx')
y_data = np.array(y_data)
x_data = np.array(x_data)

#print(y_data)
print(np.shape(x_data),np.shape(y_data))
plt.ion()
fig = plt.figure(figsize=(10,8))
ax1 = fig.add_subplot(211)
ax2 = fig.add_subplot(212)

def ROI_list(Target_array, min, max):
    Target_list = list(Target_array)
    Target_len = len(Target_list)
    min_indx = [i for i in range(0, min)]
    modifiedArray = np.delete(Target_array, min_indx,None)
    max = max - min
    min_indx = [i for i in range(max,len(modifiedArray))]
    return np.delete(modifiedArray, min_indx, None)
left = 0
right = list_len
chunk = right - left
#x_data = ROI_list(x_data,left,right)
#y_data = ROI_list(y_data,left,right)
ax1.plot(x_data,y_data)

fftData=np.fft.rfft(y_data)
fftTime=np.fft.rfftfreq(chunk, 1./1000)
#plt.plot(x_data,y_data)
print(np.shape(fftTime),np.shape(fftData))
ax2.plot(fftTime,fftData)
print(fftData)
fig.show()
plt.pause(100)