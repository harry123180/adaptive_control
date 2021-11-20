from openpyxl import load_workbook
from matplotlib import pyplot as plt
wb = load_workbook('C:\\Users\\user\\Desktop\\研究\\研究紀錄By泓舉\\9碩一上第九周\\6000RPM_GT_HU4_o.xlsx')
sheet = wb['sheet1']
Y_current = sheet['A2':'A72901']
#Y_current = load_workbook('C:\\Users\\user\\Desktop\\研究\\研究紀錄By泓舉\\8碩一上第八周\\data3.xlsx')['s1']['A2':'A9000']
print(Y_current[0][0].value)
y_data=[]
for row in Y_current:
    for c in row:
        y_data.append(c.value)
#y_data = [ int(c.value)*100 for row in Y_current for c in row]
print(y_data)
x_data = [i for i in range(0, len(y_data))]
plt.plot(x_data,y_data)
plt.show()
