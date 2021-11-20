import numpy as np
import cv2
from openpyxl import load_workbook
wb = load_workbook('C:\\Users\\user\\Desktop\\研究\\研究紀錄By泓舉\\9碩一上第九周\\6000RPM_GT_HU4_o.xlsx')
sheet = wb['sheet1']
Y_current = sheet['A2':'A72991']
#Y_current = load_workbook('C:\\Users\\user\\Desktop\\研究\\研究紀錄By泓舉\\8碩一上第八周\\data3.xlsx')['s1']['A2':'A9000']
print(Y_current[0][0].value)
y_data=[]
#創一個270*270大小array
#依序填入
img_arr = np.zeros((270,270),dtype=np.uint8)
def MaxMinNormalization(x,Max,Min):
    x = (x - Min) / (Max - Min)
    return x

for row in Y_current:
    for c in row:
        y_data.append(int(MaxMinNormalization((c.value),1.71,0)*255))


def to_matrix(l, n):
    return [l[i:i + n] for i in range(0, len(l), n)]
for i in range(270):
    for j in range(270):
        img_arr[j,i] = y_data[i*270+j]
print(img_arr)


column, row = 500, 500
img = np.zeros((column, row))
print(type(img))
re_arr = cv2.resize(img_arr,(1000,1000))
cv2.imshow("123",img_arr)
cv2.imshow("放大",re_arr)
cv2.imwrite("t4.jpg",re_arr)
cv2.waitKey(0)